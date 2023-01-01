import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# calculate how many product we have per supplier and list the name of supplier with respective number of products
products_per_supplier_dict = {}
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier_dict:
        current_num_products = products_per_supplier_dict.get(supplier_name)
        products_per_supplier_dict[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier_dict[supplier_name] = 1

        # calcualte total value of inventory per supplier
    if supplier_name in total_value_per_supplier
        total_value_per_supplier.get(supplier_name)
        total_value_per_supplier = inventory * price

